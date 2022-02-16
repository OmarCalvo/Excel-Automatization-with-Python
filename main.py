# Procesa datos en time_series_covid_19_deaths.xlsx, produciendo nuevo archivo GrafosCovidMuertes.xlsx, que
# contiene: datos originales, muertes cumulativas mundiales por día con gráfico correspondiente,
# y gráficos por país (o país y región) de muertes acumuladas diarias locales
#
# Datos descargados de:
# https://www.kaggle.com/sudalairajkumar/novel-corona-virus-2019-dataset. Muertes registradas entre 22/1/20 y 29/5/21
# Para correr, ejecutar este programa en el mismo directorio de time_series_covid_19_deaths.xlsx. Sobreescribe el
# archivo producido en ejecuciones anteriores.


from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

# Constantes
MIN_COL = 5
MAX_COL = 498
MIN_FILA = 2
MAX_FILA = 277


# Crea Grafo en base a info en fila dada de la primera hoja de trabajo
def crea_grafo_pais(fila_pais):

    grafo = LineChart()

    if ws_origen.cell(fila_pais, 1).value is None:
        grafo.title = ws_origen.cell(fila_pais, 2).value
    else:
        grafo.title = f'{ws_origen.cell(fila_pais, 2).value} - {ws_origen.cell(fila_pais, 1).value}'

    grafo.style = 13
    grafo.y_axis.title = 'Muertes'
    grafo.x_axis.title = 'Fecha'
    grafo.legend = None
    grafo.height = 12
    grafo.width = 25
    grafo.y_axis.crossAx = 500
    grafo.x_axis = DateAxis(crossAx=100)
    grafo.x_axis.number_format = 'd-mmm'
    grafo.x_axis.majorTimeUnit = "days"
    datos_acum = Reference(ws_origen, min_col=MIN_COL, min_row=fila_pais, max_col=MAX_COL, max_row=fila_pais)
    grafo.add_data(datos_acum, from_rows=True)
    grafo.set_categories(fechasAccMundiales)
    return grafo


# guardar copia nuevo y trabajar completamente sobre copia
wb_origen = load_workbook('time_series_covid_19_deaths.xlsx')
wb_origen.save('GrafosCovidMuertes.xlsx')
wb_origen.close()
wb_origen = load_workbook('GrafosCovidMuertes.xlsx')

# Recogiendo datos acumulados mundiales
ws_origen = wb_origen.active
ws_totales = wb_origen.create_sheet('TotalAcumDiario')
muertos_dia_acctotal = []
dias = []

for col in range(MIN_COL, MAX_COL + 1):
    subtotal = 0
    dias.append(ws_origen.cell(1, col).value)
    for fila in range(MIN_FILA, MAX_FILA + 1):
        subtotal += int(ws_origen.cell(fila, col).value)
    muertos_dia_acctotal.append(subtotal)

# Escribiendo acumulados mundiales en hoja totales
ws_totales.cell(1, 1).font = Font(bold=True)
ws_totales.cell(1, 1).value = 'Fecha'
ws_totales.cell(1, 2).font = Font(bold=True)
ws_totales.cell(1, 2).value = 'Total Mundial de Muertes (Acumulado)'

for i in range(0, len(dias)):
    ws_totales.cell(i + 2, 1).value = dias[i]
    ws_totales.cell(i + 2, 2).value = muertos_dia_acctotal[i]

# Graficando totales mundiales al lado de acumulados
c1 = LineChart()
c1.title = "Muertes Acumuladas Mundialmente"
c1.style = 13
c1.y_axis.title = 'Muertes'
c1.x_axis.title = 'Fecha'
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.legend = None
c1.height = 12
c1.width = 25

datosAccMundiales = Reference(ws_totales, min_col=2, min_row=2, max_col=2, max_row=495)
c1.add_data(datosAccMundiales, from_rows=False)
fechasAccMundiales = Reference(ws_totales, min_col=1, min_row=2, max_row=495)
c1.set_categories(fechasAccMundiales)
ws_totales.add_chart(c1, "G2")

# Dibujando grafos por país
ws_grafospais = None
espaciado = 25
n_grafos = 0
n_grafos_pag = 0
n_paginas = 0
fila_inicial = 2

for x in range(2, MAX_FILA+1):

    if n_grafos % 10 == 0:
        n_paginas += 1
        ws_grafospais = wb_origen.create_sheet(f'GrafosPais {n_paginas}')
        fila_inicial = 2
        n_grafos_pag = 0

    grafo_pais = crea_grafo_pais(x)
    fila_grafo = fila_inicial + n_grafos_pag * espaciado
    ws_grafospais.add_chart(grafo_pais, f'B{fila_grafo}')

    n_grafos += 1
    n_grafos_pag += 1

wb_origen.save('GrafosCovidMuertes.xlsx')
